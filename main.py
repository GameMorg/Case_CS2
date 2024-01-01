from datetime import date
import openpyxl
from openpyxl import load_workbook


def start(account, case):
    farm = 'farm.xlsx'
    # Создаем новый файл Excel
    try:
        workbook = load_workbook(filename=farm)

        # Выбираем активный лист
        sheet = workbook.active

        # Запиываем заголосвки столбцов
        sheet['A1'] = 'Имя'
        sheet['B1'] = 'Дата'
        sheet['C1'] = 'Название кейса'

        # Записываем данные в ячейки
        sheet['A2'] = 'Последний запуск'
        sheet['B2'] = date.today()

        i = 2
        while True:
            if check_value_in_row(farm, 'Sheet', i):
                set_value_in_row(sheet, 'A', str(i), account)
                set_value_in_row(sheet, 'B', str(i), date.today())
                set_value_in_row(sheet, 'C', str(i), case)
                workbook.save(farm)
                break
            else:
                i += 1

        # Сохраняем файл
        workbook.save(farm)

    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.save(farm)
        start(account, case)


def check_value_in_row(file_path, sheet_name, row_num):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    row = "B" + str(row_num)
    cell = sheet[row]
    if cell.value is None:
        return True
    return False


def set_value_in_row(sheet, colum, row_num, value):
    field = colum + row_num
    sheet[field] = value


if __name__ == '__main__':
    pass
